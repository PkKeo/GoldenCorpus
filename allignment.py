import sys

import pandas as pd

# Load the Excel file
def load_excel_column(file_path, column_name):
    """
    Load a specific column from an Excel file into a list.

    Args:
        file_path (str): Path to the Excel file.
        column_name (str): Name of the column to extract.

    Returns:
        list: A list of strings from the specified column.
    """
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Ensure the column exists
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in the Excel file.")

    # Extract the column as a list of strings
    return df[column_name].fillna("NaN").astype(str).tolist()


# Example usage
excel_file_path = "C:/Users/Admin/Desktop/Assignment/Introduction to Natural Language Processing/Final/GoldenCorpus/OCR_custom 181_182.xlsx"
column_name = "OCR_text"

# Get the list of strings from the column
ocr_texts = load_excel_column(excel_file_path, column_name)


def is_in_blacklist(string, blacklist_file):
    """
    Check if a string is in the blacklist.

    Args:
        string (str): The string to check.
        blacklist_file (str): Path to the text file containing the blacklist.

    Returns:
        bool: True if the string is in the blacklist, False otherwise.
    """
    try:
        with open(blacklist_file, 'r', encoding= 'utf-8') as file:
            # Load the blacklist into a set for fast lookup
            blacklist = {line.strip() for line in file}
        return string in blacklist
    except FileNotFoundError:
        raise FileNotFoundError(f"Blacklist file '{blacklist_file}' not found.")
    except Exception as e:
        raise Exception(f"An error occurred: {e}")


#

import os
import re
import math
import unicodedata

def normalize_word(word):
    """
    Normalize a word to lowercase and remove accents (diacritical marks).

    Args:
        word (str): The word to normalize.

    Returns:
        str: The normalized word.
    """
    # Convert to lowercase
    word = word.lower()
    # Remove accents by decomposing the Unicode characters and filtering
    word = ''.join(
        char for char in unicodedata.normalize('NFD', word)
        if unicodedata.category(char) != 'Mn'
    )
    word = re.sub(r'-', '', word)
    word = re.sub(r'f', 't', word)
    word = re.sub(r'j', 'i', word)
    word = re.sub(r'1', 'i', word)
    word = re.sub(r'4', 'a', word)
    word = re.sub(r'7', '?', word)
    return word


def normalize_text(text):
    # Strip whitespace and normalize hyphenation and punctuation rules
    text = text.strip()

    # Ensure words like 'bê-tông' are split as 'bê - tông'
    text = re.sub(r'(\w)-(\w)', r'\1 - \2', text)  # Insert a space around the hyphen between words
    text = re.sub(r'\s+t\s+', ':', text)
    text = re.sub(r'\s+t$', ':', text)
    text = re.sub(r'–', '-', text)
    text = re.sub(r'^\?\s+', '?', text)

    text = re.sub(r'-\s+([A-Z])', r'-\1', text)
    text = re.sub(r'\.\.\.', '…', text)
    text = re.sub(r'\s+7\s+', '?', text)
    text = re.sub(r'\s+7$', '?', text)

    # Remove spaces before punctuation
    text = re.sub(r'\s([.,;:?!])', r'\1', text)

    # Ensure punctuation is followed by a space
    text = re.sub(r'([.,;:?!])\s', r'\1 ', text)


    text = text.strip()

    return text


def natural_sort_key(filename):
    """
    Generate a sorting key that extracts numbers from the filename for natural sorting.
    Example: 'page10.txt' > 10, 'page2.txt' > 2.
    """
    match = re.search(r'\d+', filename)
    return int(match.group()) if match else float('inf')  # Sort non-matching filenames last

def read_files_in_folder(folder, skip=0):
    """
    Read all text files in the folder, sorted by their numeric filename order,
    and optionally skip the first `skip` files.

    Args:
        folder (str): Path to the folder containing the files.
        skip (int): Number of files to skip at the start.

    Yields:
        str: Contents of each file in order, after skipping the specified number.
    """
    # List all .txt files in the folder and sort them using the natural sorting key
    files = sorted([f for f in os.listdir(folder) if f.endswith('.txt')], key=natural_sort_key)

    # Skip the specified number of files
    files_to_process = files[skip:]

    # Read each remaining file in the sorted order
    for file in files_to_process:
        with open(os.path.join(folder, file), 'r', encoding='utf-8') as f:
            yield normalize_text(f.read())


import difflib

def levenshtein_distance(a, b):
    """Calculate the Levenshtein distance between two strings."""
    len_a, len_b = len(a), len(b)
    dp = [[0] * (len_b + 1) for _ in range(len_a + 1)]

    # Initialize base cases
    for i in range(len_a + 1):
        dp[i][0] = i  # Cost of deleting all characters from `a`
    for j in range(len_b + 1):
        dp[0][j] = j  # Cost of inserting all characters into `a`

    # Compute distances
    for i in range(1, len_a + 1):
        for j in range(1, len_b + 1):
            if a[i - 1] == b[j - 1]:
                cost = 0  # No cost if characters match
            else:
                cost = 1  # Substitution cost
            dp[i][j] = min(
                dp[i - 1][j] + 1,      # Deletion
                dp[i][j - 1] + 1,      # Insertion
                dp[i - 1][j - 1] + cost  # Substitution
            )

    return dp[len_a][len_b]


def words_similar(word1, word2, threshold=0.5):
    """
    Check if two words are similar based on Levenshtein distance
    after normalization.

    Args:
        word1 (str): First word.
        word2 (str): Second word.
        threshold (float): Maximum allowed distance as a fraction of word length.

    Returns:
        bool: True if the words are similar enough, False otherwise.
    """
    # Normalize both words
    word1 = normalize_word(word1)
    word2 = normalize_word(word2)
    # Calculate Levenshtein distance
    distance = levenshtein_distance(word1, word2)
    if(distance <= threshold * max(len(word2), len(word1))):
      return True , distance
    else:
      return False , distance



def match_and_reduce_with_alignment(typo_sentence, buffer, similarity_threshold=0.5):
    """
    Match the typo sentence with the buffer and reduce the matched part,
    returning the matched segment and the reduced buffer.
    """
    typo_words = typo_sentence.split()
    buffer_words = buffer.split()
    n = len(typo_words)

    for i in range(len(buffer_words) - n + 1):  # Iterate through buffer with sliding window
        for j in [n]:  # Extend range to include up to one extra word
            if i + j > len(buffer_words):
                break  # Avoid index out of range
            # if(j == n-1):
            #   print('Here')
            # Check word-by-word similarity
            segment = buffer_words[i:i + j]
            # if len(segment) != len(typo_words):
            #     continue
            if len(segment) == len(typo_words):
              similarity_count = 0
              for typo_word, segment_word in zip(typo_words, segment):
                res, dis =(words_similar(typo_word, segment_word, similarity_threshold))
                #print(f"res = {res}")
                if(res):
                  similarity_count += 1
              #print(similarity_count)

              # If all words are similar enough, return the matched segment and reduced buffer
              if similarity_count >= math.ceil(len(typo_words) * 1):
                  matched_segment = " ".join(segment)
                  reduced_buffer = " ".join(buffer_words[i + j:])
                  #print("This")
                  return matched_segment, reduced_buffer, True


              buffer_sentence = "".join(buffer_words[i:i + j])
              typo_sentence = "".join(typo_words)
              buffer_sentence = normalize_text(buffer_sentence)
              typo_sentence = normalize_text(typo_sentence)
              if(words_similar(buffer_sentence, typo_sentence, 0.1)[0]):
                matched_segment = " ".join(segment)
                reduced_buffer = " ".join(buffer_words[i + j:])
                # print(buffer_sentence)
                # print(typo_sentence)

                # print("that")
                return matched_segment, reduced_buffer, True


    best_match_segment = None
    best_match_distance = float('inf')
    best_reduce_buffer = buffer
    for i in range(min(len(buffer_words) - n + 1, 20)):  # Iterate through buffer with sliding window
        for j in [n + 1, n - 1, n + 2, n - 2]:  # Extend range to include up to one extra word
            if i + j > len(buffer_words):
                break  # Avoid index out of range
            # Check word-by-word similarity
            segment = buffer_words[i:i + j]
            # if len(segment) != len(typo_words):
            #     continue
            buffer_sentence = "".join(buffer_words[i:i + j])
            typo_sentence = "".join(typo_words)
            buffer_sentence = normalize_text(buffer_sentence)
            res, dis = words_similar(buffer_sentence, typo_sentence, 0.2)
            if(res):
              matched_segment = " ".join(segment)
              reduced_buffer = " ".join(buffer_words[i + j:])
              if(dis < best_match_distance):
                best_match_distance = dis
                best_match_segment = matched_segment
                best_reduce_buffer = reduced_buffer
    if(best_match_segment):
      #print("Noise")
      return best_match_segment, best_reduce_buffer, False

    return None, buffer, False  # No suitable alignment found

def does_not_contain_vietnamese(text):
    # Match any Vietnamese alphabet character (including accented characters)
    vietnamese_alphabet = r'[aăâáàảãạắằẳẵặấầẩẫậeêéèẻẽẹếềểễệiíìỉĩịoôơóòỏõọốồổỗộớờởỡợuưúùủũụứừửữựyýỳỷỹỵđ]'
    # If no match is found, the string does not contain Vietnamese alphabet
    return not re.search(vietnamese_alphabet, text, re.IGNORECASE)

def process_typo_sentences_with_alignment(folder, list_typo_sentence):
    buffer = ""
    file_iterator = read_files_in_folder(folder, skip  = 33 )  # Initialize file generator
    alignments = []  # To store alignments
    #alignment have 2 part: string and type
    #type 0: correct in number of words with OCR
    #type 1: not correct in number of words with OCR
    #type 2: This is skipped during work
    #type 3: This is not a good OCR (wrong text detection)
    count = 2
    patience = 0
    blacklist_file = r"C:\Users\Admin\Desktop\Assignment\Introduction to Natural Language Processing\Final\GoldenCorpus\black_list.txt"
    for typo_sentence in list_typo_sentence:
        print(f'Processing on sentence {count}')
        if(does_not_contain_vietnamese(typo_sentence) or is_in_blacklist(typo_sentence, blacklist_file)):
          count += 1
          alignments.append(('NOT A WORD', 3))
          continue
        typo_sentence = normalize_text(typo_sentence)
        match_found = False  # Flag to indicate if a match was found
        cnt = 0
        skip = False
        while not match_found:
            # If buffer is empty, load more text
            if not buffer:
                try:
                    buffer += " " + next(file_iterator)
                    buffer = normalize_text(buffer)
                except StopIteration:
                    break  # No more files to read

            # Process the typo sentence
            matched_segment, new_buffer, same_len = match_and_reduce_with_alignment(typo_sentence, buffer)
            if matched_segment:  # Match found
                # print(count)
                # print(f"Typo Sentence: {typo_sentence}")
                # print(f"Matched: {matched_segment}")
                if(same_len):
                    alignments.append((matched_segment, 0))
                else:
                    alignments.append((matched_segment, 1))
                buffer = new_buffer  # Update buffer
                patience = 0
                match_found = True  # Set flag to True and exit the while loop
            else:  # No match yet, load more text
                print(f"No match found for Typo Sentence: {normalize_text(typo_sentence)}")
                print(f"buffer: {normalize_text(buffer)}")
                if(cnt >= 1):
                    skip = True
                    patience += 1
                    break
                try:
                    cnt += 1
                    buffer += " " + next(file_iterator)
                    buffer = normalize_text(buffer)
                except StopIteration:
                    break  # No more files to read

        if skip:
          # print(f"SKIP found for Typo Sentence: {normalize_text(typo_sentence)}")
          alignments.append(("SKIPPED", 2))
          count += 1
          continue
        if(patience >= 2):
          sys.exit()

        # if not match_found:  # If no match was found after reading all files
        #     print(f"No match found for Typo Sentence: {normalize_text(typo_sentence)}")
        #     print(f"buffer: {normalize_text(buffer)}")

        count += 1

    return alignments



# Example usage
folder_path = r"C:\Users\Admin\Desktop\Assignment\Introduction to Natural Language Processing\Final\GoldenCorpus\extracted_text\content\vietnamese-ocr\extracted_text"
alignments = process_typo_sentences_with_alignment(folder_path, ocr_texts)


from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def add_correct_text_column(file_path, alignments):
    """
    Add a new column named 'Correct Text' to an existing Excel file and write the alignments list with formatting.

    Args:
        file_path (str): Path to the existing Excel file.
        alignments (list): List of tuples (string, int) to write.
    """
    # Define color mapping
    color_map = {
        0: "000000",  # Black
        1: "00FF00",  # Green
        2: "0000FF",  # Blue
        3: "FF0000",  # Red
    }

    # Load the existing workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Check if the "Correct Text" column already exists
    correct_text_column = None
    for col_idx, cell in enumerate(sheet[1], start=1):  # Iterate through the first row
        if cell.value == "Correct Text":
            correct_text_column = col_idx
            break

    # If the column does not exist, add it
    if correct_text_column is None:
        correct_text_column = 7
        sheet.cell(row=1, column=correct_text_column, value="Correct Text")

    # Write the alignments with formatting
    for i, (text, color_code) in enumerate(alignments, start=2):  # Start writing from the second row
        cell = sheet.cell(row=i, column=correct_text_column, value=text)

        # Apply font color
        cell.font = Font(color=color_map.get(color_code, "000000"))

    # Save the workbook
    workbook.save(file_path)


add_correct_text_column(excel_file_path, alignments)
