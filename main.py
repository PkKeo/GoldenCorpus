import sys
import os
import re
import math
import pandas as pd
import unicodedata
from dataclasses import dataclass
from typing import List, Tuple, Iterator, Optional
from enum import IntEnum
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from typing import Union

class AlignmentType(IntEnum):
    """Enumeration for different types of alignments"""
    EXACT_MATCH = 0  # Correct number of words with OCR
    PARTIAL_MATCH = 1  # Different number of words than OCR
    SKIPPED = 2  # Skipped during processing
    INVALID = 3  # Not a good OCR (wrong text detection)

@dataclass
class AlignmentResult:
    """Data class to store alignment results"""
    text: str
    type: AlignmentType

class TextNormalizer:
    """Class for text normalization operations"""

    @staticmethod
    def normalize_word(word: str) -> str:
        """Normalize a single word by removing accents and applying substitutions"""
        word = word.lower()
        word = ''.join(
            char for char in unicodedata.normalize('NFD', word)
            if unicodedata.category(char) != 'Mn'
        )

        substitutions = {
            'f': 't',
            'j': 'i',
            '1': 'i',
            '4': 'a',
            '7': '?',
            '-': ''
        }

        for old, new in substitutions.items():
            word = word.replace(old, new)
        return word

    @staticmethod
    def normalize_text(text: str) -> str:
        """Normalize text by applying various cleaning rules"""
        text = text.strip()

        # Define regex patterns and their replacements
        patterns = [
            (r'(\w)-(\w)', r'\1 - \2'),  # Split hyphenated words
            (r'\s+t\s+|\s+t$', ':'),     # Replace 't' with ':'
            (r'–', '-'),                  # Standardize hyphens
            (r'^\?\s+', '?'),            # Clean up question marks
            (r'-\s+([A-Z])', r'-\1'),    # Fix spacing around hyphens
            (r'\.\.\.', '…'),            # Replace triple dots with ellipsis
            (r'\s+7\s+|\s+7$', '?'),     # Replace '7' with '?'
            (r'\s([.,;:?!])', r'\1'),    # Remove spaces before punctuation
            (r'([.,;:?!])\s', r'\1 ')    # Ensure space after punctuation
        ]

        for pattern, replacement in patterns:
            text = re.sub(pattern, replacement, text)

        return text.strip()

class LevenshteinCalculator:
    """Class for calculating Levenshtein distance and word similarity"""

    @staticmethod
    def distance(a: str, b: str) -> int:
        """Calculate the Levenshtein distance between two strings"""
        if len(a) < len(b):
            return LevenshteinCalculator.distance(b, a)

        if not b:
            return len(a)

        previous_row = range(len(b) + 1)
        for i, c1 in enumerate(a):
            current_row = [i + 1]
            for j, c2 in enumerate(b):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    @staticmethod
    def words_similar(word1: str, word2: str, threshold: float = 0.5) -> Tuple[bool, int]:
        """Check if two words are similar based on Levenshtein distance"""
        word1 = TextNormalizer.normalize_word(word1)
        word2 = TextNormalizer.normalize_word(word2)
        distance = LevenshteinCalculator.distance(word1, word2)
        is_similar = distance <= threshold * max(len(word2), len(word1))
        return is_similar, distance

class TextMatcher:
    """Class for matching and aligning text segments"""

    def __init__(self, similarity_threshold: float = 0.5):
        self.similarity_threshold = similarity_threshold

    def match_and_reduce(self, typo_sentence: str, buffer: str) -> Tuple[Optional[str], str, bool]:
        """Match typo sentence with buffer and reduce matched part"""
        typo_words = typo_sentence.split()
        buffer_words = buffer.split()
        n = len(typo_words)

        # Try exact match first
        exact_match = self._try_exact_match(typo_words, buffer_words, n)
        if exact_match:
            return exact_match

        # Try fuzzy match if exact match fails
        return self._try_fuzzy_match(typo_words, buffer_words, n)

    def _try_exact_match(self, typo_words: List[str], buffer_words: List[str], n: int) -> Optional[Tuple[str, str, bool]]:
        """Attempt to find an exact match"""
        for i in range(len(buffer_words) - n + 1):
            segment = buffer_words[i:i + n]
            if len(segment) == len(typo_words):
                if self._check_similarity(typo_words, segment):
                    return self._create_match_result(segment, buffer_words[i + n:], True)
        return None

    def _try_fuzzy_match(self, typo_words: List[str], buffer_words: List[str], n: int) -> Tuple[Optional[str], str, bool]:
        """Attempt to find a fuzzy match"""
        best_match = {
            'segment': None,
            'distance': float('inf'),
            'reduced_buffer': buffer_words
        }

        for i in range(min(len(buffer_words) - n + 1, 20)):
            for j in [n + 1, n - 1, n + 2, n - 2]:
                if i + j > len(buffer_words):
                    continue

                match = self._check_fuzzy_match(
                    typo_words,
                    buffer_words[i:i + j],
                    buffer_words[i + j:],
                    best_match
                )

                if match:
                    best_match = match

        if best_match['segment']:
            return self._create_match_result(
                best_match['segment'],
                best_match['reduced_buffer'],
                False
            )

        return None, ' '.join(buffer_words), False

    def _check_similarity(self, typo_words: List[str], segment: List[str]) -> bool:
        """Check if words are similar enough"""
        similarity_count = sum(
            1 for t, s in zip(typo_words, segment)
            if LevenshteinCalculator.words_similar(t, s, self.similarity_threshold)[0]
        )
        return similarity_count >= math.ceil(len(typo_words))

    def _check_fuzzy_match(
        self,
        typo_words: List[str],
        segment: List[str],
        remaining_buffer: List[str],
        best_match: dict
    ) -> Optional[dict]:
        """Check for fuzzy matches and update best match if found"""
        buffer_sentence = TextNormalizer.normalize_text(''.join(segment))
        typo_sentence = TextNormalizer.normalize_text(''.join(typo_words))

        is_similar, distance = LevenshteinCalculator.words_similar(
            buffer_sentence,
            typo_sentence,
            0.2
        )

        if is_similar and distance < best_match['distance']:
            return {
                'segment': segment,
                'distance': distance,
                'reduced_buffer': remaining_buffer
            }
        return None

    @staticmethod
    def _create_match_result(
        segment: List[str],
        remaining_buffer: List[str],
        exact_match: bool
    ) -> Tuple[str, str, bool]:
        """Create a match result tuple"""
        return (
            ' '.join(segment),
            ' '.join(remaining_buffer),
            exact_match
        )

class TextProcessor:
    """Main class for processing and aligning text"""

    def __init__(self, folder_path: str, blacklist_path: str):
        self.folder_path = Path(folder_path)
        self.blacklist_path = Path(blacklist_path)
        self.matcher = TextMatcher()
        self._load_blacklist()
        self.patience = 0
        self.max_patience = 2  # Maximum number of consecutive skips before exiting

    def _load_blacklist(self) -> None:
        """Load blacklist from file"""
        try:
            with open(self.blacklist_path, 'r', encoding='utf-8') as file:
                self.blacklist = {line.strip() for line in file}
        except FileNotFoundError:
            raise FileNotFoundError(f"Blacklist file '{self.blacklist_path}' not found.")

    def process_texts(self, ocr_texts: List[str], skip_files: int = 33) -> List[AlignmentResult]:
        """Process OCR texts and return alignments"""
        buffer = ""
        alignments = []
        file_iterator = self._read_files(skip_files)

        for i, text in enumerate(ocr_texts, start=2):
            print(f'Processing sentence {i}')

            if self._should_skip(text):
                alignments.append(AlignmentResult("NOT A WORD", AlignmentType.INVALID))
                continue

            result = self._process_single_text(text, buffer, file_iterator)
            if isinstance(result, AlignmentResult):
                alignments.append(result)
            else:
                alignment, new_buffer = result
                alignments.append(alignment)
                buffer = new_buffer

            if self.patience >= self.max_patience:
                print("Maximum patience exceeded. Exiting...")
                sys.exit(1)

        return alignments

    def _process_single_text(self, text: str, buffer: str, file_iterator: Iterator[str]) -> Union[AlignmentResult, Tuple[AlignmentResult, str]]:
        """Process a single text entry and return alignment result with updated buffer"""
        text = TextNormalizer.normalize_text(text)
        match_found = False
        attempts = 0
        max_attempts = 2
        current_buffer = buffer

        while not match_found and attempts < max_attempts:
            # Load more text if buffer is empty
            if not current_buffer:
                try:
                    current_buffer = TextNormalizer.normalize_text(next(file_iterator))
                except StopIteration:
                    break

            # Try to match the text
            matched_segment, new_buffer, exact_match = self.matcher.match_and_reduce(text, current_buffer)

            if matched_segment:
                self.patience = 0  # Reset patience counter on successful match
                alignment_type = AlignmentType.EXACT_MATCH if exact_match else AlignmentType.PARTIAL_MATCH
                return AlignmentResult(matched_segment, alignment_type), new_buffer

            # No match found, try loading more text
            if attempts < max_attempts - 1:
                try:
                    next_text = next(file_iterator)
                    current_buffer += " " + TextNormalizer.normalize_text(next_text)
                except StopIteration:
                    break

            attempts += 1
            print(f"No match found for text: {text}")
            print(f"Current buffer: {current_buffer}")

        # If we reach here, no match was found
        self.patience += 1
        return AlignmentResult("SKIPPED", AlignmentType.SKIPPED)

    def _should_skip(self, text: str) -> bool:
        """Determine if text should be skipped"""
        return (
            self._contains_no_vietnamese(text) or
            text in self.blacklist
        )

    @staticmethod
    def _contains_no_vietnamese(text: str) -> bool:
        """Check if text contains no Vietnamese characters"""
        vietnamese_pattern = r'[aăâáàảãạắằẳẵặấầẩẫậeêéèẻẽẹếềểễệiíìỉĩịoôơóòỏõọốồổỗộớờởỡợuưúùủũụứừửữựyýỳỷỹỵđ]'
        return not re.search(vietnamese_pattern, text, re.IGNORECASE)

    def _read_files(self, skip_files: int) -> Iterator[str]:
        """Generate normalized text from files"""
        # Debug print to show all available files
        all_files = [f for f in self.folder_path.glob('*.txt')]
        print(f"\nTotal number of files found: {len(all_files)}")

        # Sort files and print first few file names
        files = sorted(
            all_files,
            key=lambda x: int(re.search(r'\d+', x.name).group())
        )
        print("\nFirst 5 files before skip:")
        for f in files[:5]:
            print(f"- {f.name}")

        # Print skip information
        print(f"\nSkipping {skip_files} files")

        # Get files after skip
        files_after_skip = files[skip_files:]
        print("\nFirst 5 files after skip:")
        for f in files_after_skip[:5]:
            print(f"- {f.name}")

        # Yield file contents
        for file in files_after_skip:
            print(f"\nReading file: {file.name}")
            with open(file, 'r', encoding='utf-8') as f:
                content = TextNormalizer.normalize_text(f.read())
                print(f"First 100 characters: {content[:100]}...")
                yield content

class ExcelWriter:
    """Class for writing results to Excel"""

    COLOR_MAP = {
        AlignmentType.EXACT_MATCH: "000000",  # Black
        AlignmentType.PARTIAL_MATCH: "00FF00", # Green
        AlignmentType.SKIPPED: "0000FF",      # Blue
        AlignmentType.INVALID: "FF0000"       # Red
    }

    def __init__(self, file_path: str):
        self.file_path = file_path

    def write_alignments(self, alignments: List[AlignmentResult]) -> None:
        """Write alignment results to Excel file"""
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        correct_text_column = self._get_or_create_column(sheet)
        self._write_results(sheet, alignments, correct_text_column)

        workbook.save(self.file_path)

    def _get_or_create_column(self, sheet) -> int:
        """Get existing or create new 'Correct Text' column"""
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Correct Text":
                return idx

        column = 7
        sheet.cell(row=1, column=column, value="Correct Text")
        return column

    def _write_results(self, sheet, alignments: List[AlignmentResult], column: int) -> None:
        """Write results with formatting"""
        for i, alignment in enumerate(alignments, start=2):
            cell = sheet.cell(row=i, column=column, value=alignment.text)
            cell.font = Font(color=self.COLOR_MAP[alignment.type])

def main():
    """Main execution function"""
    # Configuration
    excel_file = "C:/Users/Admin/Desktop/Assignment/Introduction to Natural Language Processing/Final/GoldenCorpus/OCR_custom 181_182.xlsx"
    folder_path = "C:/Users/Admin/Desktop/Assignment/Introduction to Natural Language Processing/Final/GoldenCorpus/extracted_text/content/vietnamese-ocr/extracted_text"
    blacklist_file = "C:/Users/Admin/Desktop/Assignment/Introduction to Natural Language Processing/Final/GoldenCorpus/black_list.txt"

    # Debug print current configuration
    print("\nConfiguration:")
    print(f"Excel file: {excel_file}")
    print(f"Folder path: {folder_path}")
    print(f"Blacklist file: {blacklist_file}")

    try:
        # Load OCR texts
        df = pd.read_excel(excel_file)
        ocr_texts = df["OCR_text"].fillna("NaN").astype(str).tolist()
        print(f"\nNumber of OCR texts to process: {len(ocr_texts)}")
        print("First few OCR texts:")
        for i, text in enumerate(ocr_texts[:3], 1):
            print(f"{i}. {text[:100]}...")

        # Set which page to start from
        start_page = 193
        # Print the file numbers we're looking at
        print(f"\nTarget start page: {start_page}")

        # Process texts with specified starting page
        processor = TextProcessor(folder_path, blacklist_file)

        # Get first file number to calculate skip
        first_file = sorted([f for f in Path(folder_path).glob('*.txt')])[0]
        first_page_num = int(re.search(r'\d+', first_file.name).group())
        files_to_skip = start_page - first_page_num
        print(f"First file starts at page: {first_page_num}")
        print(f"Files to skip: {files_to_skip}")

        alignments = processor.process_texts(ocr_texts, skip_files=files_to_skip)

        # Write results
        writer = ExcelWriter(excel_file)
        writer.write_alignments(alignments)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
